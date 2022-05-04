# 将wrf输出值插值到需要的格点上

from lib.Interpolate import interpolate,calculate_wind_direction
import lib.Readtime as Readtime
from wrf import getvar,to_np
from openpyxl import Workbook
import numpy as np
import netCDF4 as nc

#下面修改路径
path=r'D:\Data\WRF-Chem_Files\WRF-Chem_Simulation\lcz_8\wrfout_d03_2016-07-21'

#下面修改站点的纬度，经度，气象站点名字。气象站点名字可以带中文，主要就是为了excel表格的sheet书写
#point_list=[(31.1,121.37,"58361闵行"),(31.39692,121.45454,"58362宝山"),(31.37,121.25,"58365嘉定"),(31.67,121.50,"58366崇明"),
#            (31.05,121.7833,"58369南汇"),(31.13,121.12,"58461青浦"),(30.88,121.50,"58463奉贤"),(30.73,121.35,"58460金山")]
point_list=[(31.1,121.3667,"58361闵行"),(31.4,121.45,"58362宝山"),(31.3667,121.25,"58365嘉定"),
            (31.05,121.7833,"58369南汇"),(30.8833,121.50,"58463奉贤"),(31.2,121.4333,"58367徐家汇"),(31.2333,121.5333,"58370浦东"),(30.7333,121.35,"58460金山")]
ncfile=nc.Dataset(path)
wb=Workbook()
for i in point_list:
    wb.create_sheet(str(i[2]))
    ws = wb[str(i[2])]
    ws.cell(1,1,"时间")
    ws.cell(1,2,"2m温度")
    ws.cell(1,3,"2m湿度")
    ws.cell(1,4,"10m风向")
    ws.cell(1,5,"10m风速")
    ws.cell(1,6,"u")
    ws.cell(1,7,"v")
    ws.cell(1,8,"LH")
    ws.cell(1,9,"HFX")
    ws.cell(1,10,"GRDFLX")

timelist=Readtime.get_ncfile_time(ncfile)
timestep=2
for i in range(0,Readtime.get_ncfile_alltime(ncfile),timestep):
    t2=to_np(getvar(ncfile,"T2",timeidx=i))-273.15
    rh2=to_np(getvar(ncfile,'rh2',timeidx=i))
    u10 = to_np(getvar(ncfile, "U10", timeidx=i))
    v10 = to_np(getvar(ncfile,"V10",timeidx=i))
    lh = to_np(getvar(ncfile,"LH",timeidx=i))
    hfx = to_np(getvar(ncfile,"HFX",timeidx=i))
    grdflx = to_np(getvar(ncfile,"GRDFLX",timeidx=i))
    for j in range(len(point_list)):
        float_t2 = interpolate(ncfile, t2, point_list[j][0], point_list[j][1], opt=0)
        float_rh2 = interpolate(ncfile, rh2, point_list[j][0], point_list[j][1], opt=0)
        float_u10 = interpolate(ncfile, u10, point_list[j][0], point_list[j][1], opt=0)
        float_v10 = interpolate(ncfile, v10, point_list[j][0], point_list[j][1], opt=0)
        float_lh = interpolate(ncfile, lh, point_list[j][0], point_list[j][1], opt=0)
        float_hfx = interpolate(ncfile, hfx, point_list[j][0], point_list[j][1], opt=0)
        float_grdflx = interpolate(ncfile, grdflx, point_list[j][0], point_list[j][1], opt=0)
        print("u10,v10:" + str(float_u10) + "," + str(float_v10))
        float_ws = np.sqrt(float_v10**2+float_u10**2)
        float_wdir = calculate_wind_direction(float_u10,float_v10)
        print("当前时间为：{}".format(timelist[i]))
        #print("u10,v10:"+str(float_u10)+","+str(float_v10)+","+str(float_ws)+","+str(float_wdir))
        print("行数为：{}".format(int(i/timestep+2)))
        worksheet=wb[str(point_list[j][2])]
        #wb[str(point_list[j][2])].cell(i/timestep+2,1).value = timelist[i]
        worksheet.cell(int(i/timestep+2),1,timelist[i])
        worksheet.cell(int(i/timestep+2),2,float_t2)
        worksheet.cell(int(i/timestep+2),3,float_rh2)
        worksheet.cell(int(i/timestep+2),4,float_wdir)
        worksheet.cell(int(i/timestep+2),5,float_ws)
        worksheet.cell(int(i/timestep+2),6,float_u10)
        worksheet.cell(int(i/timestep+2),7,float_v10)
        worksheet.cell(int(i/timestep+2),8,float_lh)
        worksheet.cell(int(i/timestep+2),9,float_hfx)
        worksheet.cell(int(i/timestep+2),10,float_grdflx)
wb.remove(wb['Sheet'])
wb.save("气象信息-站点插值-lcz8.xlsx")




